import openpyxl
import string
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, Font
from datetime import date

cell_border = Border (left=Side (border_style='thin', color='FF000000'),
                      right=Side (border_style='thin', color='FF000000'),
                      top=Side (border_style='thin', color='FF000000'),
                      bottom=Side (border_style='thin', color='FF000000'))

fecha = date.today ()
year = fecha.year

def ValoresCeldas(df, open_file=flow_variables['Macro'], save_file=flow_variables['Macro'], input_table_num=1,
                  get_row=4,row_size_empty=3, get_col=15, size_col=4, get_sheet="Clúster_x_seg"):



    """


    :param df: tabla de entrada().
    :param open_file: ruta del archivo a abrir(predeterminado=flow_variables['Macro']).
    :param save_file: ruta del archivo a guardar(predeterminado=flow_variables['Macro']).
    :param input_table_num: numero de tabla que entra(input_table_1).
    :param get_row: obtener la posicion de la fila(predeterminado=4).
    :param row_size_empty: tamaño predeterminado para las tablas vacías(predeterminado=3).
    :param get_col: obtener la posicion de la columna(predeterminado=15).
    :param size_col: número de columnas a tomar(predeterminado=4).
    :param get_sheet: obtener la pestaña del excel(predeterminado=Clúster_x_seg).
    :return:
    """


    if df.empty:
        print ("Tabla Vacía ", input_table_num)
        wb = openpyxl.load_workbook (open_file, read_only=False, keep_vba=True)  # Abrir archivo como macro.
        ws = wb.active
        ws = wb[get_sheet]

        ##Generar línea del cuadrado.
        for col in range (get_col, get_col + size_col):
            for row in range (get_row, get_row + row_size_empty):
                cell = ws.cell (row, col +1)
                cell.border = cell_border

        print (f"Se ha escrito, el cuadrado en la tabla {input_table_num}")

        wb.save (save_file)  # guardar archivo como macro.
        wb.close () # cierra conexión.


    else:
        wb = openpyxl.load_workbook (open_file, read_only=False, keep_vba=True)  # Abrir archivo como macro.
        ws = wb.active # activa el archivo.
        ws = wb[get_sheet] # obtiene el nombre de la pestaña.

        data = [i for i in dataframe_to_rows (df, index=False, header=True)]  # conversion de df a lista.

        celda_abc = list (string.ascii_uppercase)  # lista de alfabeto.
        celda_abc = celda_abc[get_col:] # almacena el tamaño de las columunas.

        # desemaquetar datos

        cord_data = []
        for eje_x in range (0, df.shape[0] + 1):
            for eje_y in range (0, len (data[1])):
                cord_data.append ([eje_x, eje_y])

        ##Calcular el alfabeto
        cord_exc = []
        for num_letra in range (get_row, get_row + df.shape[0] + 1):  # movimientos en el excel
            for letra_abc in range (0, len (data[1])):
                cord_exc.append (celda_abc[letra_abc] + str (num_letra))

        ##Generar cuadro
        for col in range (get_col, get_col + size_col):
            for row in range (get_row, get_row + df.shape[0] + 1):  # movimientos en el excel
                cell = ws.cell (row, col + 1)
                cell.border = cell_border

        print (cord_data)
        print ("______________________________________________________")
        print (cord_exc)
        print ("______________________________________________________")
        cord_abc_data = dict (zip (cord_exc, cord_data))
        print (cord_abc_data)
        print ("______________________________________________________")
        print (data)


        for cord_exc, x in cord_abc_data.items ():
            ws[cord_exc] = data[x[0]][x[1]]

        print (f"tabla {input_table_num}, Celda {get_row} ")
        wb.save (save_file)  # guardar archivo como macro.
        wb.close ()


ValoresCeldas (input_table_1, input_table_num=1, row_size_empty=4 + 2)  #rows arranca 4= 12+3
ValoresCeldas (input_table_2, input_table_num=2, get_row=15, row_size_empty=4 + 2)  # 4 +2 rows = 24
ValoresCeldas (input_table_3, input_table_num=3, get_row=24, row_size_empty=4 + 2)  # 4 +2 rows = 33
ValoresCeldas (input_table_4, input_table_num=4, get_row=33, row_size_empty=3 + 2)  # 3 +2 rows = 41
ValoresCeldas (input_table_5, input_table_num=5, get_row=41, row_size_empty=3 + 2)  # 3 +2 rows = 49
ValoresCeldas (input_table_6, input_table_num=6, get_row=49, row_size_empty=6 + 2)  # 6 +2 rows = 60
ValoresCeldas (input_table_7, input_table_num=7, get_row=60, row_size_empty=6 + 2)  # 6 +2 rows = 71
ValoresCeldas (input_table_8, input_table_num=8, get_row=71, row_size_empty=5 + 2)  # 5 +2 rows = 81
ValoresCeldas (input_table_9, input_table_num=9, get_row=81, row_size_empty=5 + 2)  # 5 +2 rows = 91
ValoresCeldas (input_table_10, input_table_num=10, get_row=91, row_size_empty=6 + 2)  # 6  +2 rows = 102
ValoresCeldas (input_table_11, input_table_num=11, get_row=102, row_size_empty=5 + 2)  # 5 +2 rows = 112
ValoresCeldas (input_table_12, input_table_num=12, get_row=112, row_size_empty=7 + 2)  # 7 +2 rows


